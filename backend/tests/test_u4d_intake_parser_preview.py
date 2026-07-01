"""U4-D intake parser and staging preview tests."""
from __future__ import annotations

import io
from pathlib import Path

import pytest
from fastapi import HTTPException
from openpyxl import Workbook

from services.intake_service import IntakeService
from tests.test_u4c_intake_api_contract import _client_for, _ensure_intake_schema


BACKEND_DIR = Path(__file__).resolve().parents[1]


def _xlsx_bytes(headers: list[str], rows: list[list[str]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _assert_parse_error(filename: str, file_bytes: bytes, expected_code: str) -> None:
    with pytest.raises(HTTPException) as exc_info:
        IntakeService().parse_upload(filename=filename, content_type=None, file_bytes=file_bytes)

    assert exc_info.value.status_code == 400
    assert exc_info.value.detail["code"] == expected_code


def test_parser_accepts_utf8_sig_csv_and_preserves_source_rows():
    service = IntakeService()
    payload = "\ufeffSku Code,Name,Unit\n\nSKU-1,Widget,piece\nSKU-2,Gadget,box\n".encode("utf-8")

    parsed = service.parse_upload(filename="products.csv", content_type="text/csv", file_bytes=payload)

    assert parsed.file_ext == "csv"
    assert parsed.headers_raw == ["Sku Code", "Name", "Unit"]
    assert parsed.headers_normalized == {"Sku Code": "sku_code", "Name": "name", "Unit": "unit"}
    assert parsed.source_row_numbers == [3, 4]
    assert parsed.rows[0] == {"Sku Code": "SKU-1", "Name": "Widget", "Unit": "piece"}


def test_parser_makes_duplicate_headers_deterministic():
    parsed = IntakeService().parse_upload(
        filename="products.csv",
        content_type="text/csv",
        file_bytes=b"Name,Name\nPrimary,Secondary\n",
    )

    assert parsed.headers_raw == ["Name", "Name"]
    assert parsed.headers_normalized == {"Name": "name", "Name__2": "name_2"}
    assert parsed.rows == [{"Name": "Primary", "Name__2": "Secondary"}]


def test_parser_uses_first_non_empty_xlsx_sheet():
    workbook = Workbook()
    empty_sheet = workbook.active
    empty_sheet.title = "Empty"
    data_sheet = workbook.create_sheet("Products")
    data_sheet.append(["Sku Code", "Name"])
    data_sheet.append(["XLS-1", "XLS Product"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()

    parsed = IntakeService().parse_upload(
        filename="products.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        file_bytes=buffer.getvalue(),
    )

    assert parsed.file_ext == "xlsx"
    assert parsed.parse_summary["sheet_name"] == "Products"
    assert parsed.source_row_numbers == [2]
    assert parsed.rows == [{"Sku Code": "XLS-1", "Name": "XLS Product"}]


def test_parser_rejects_unreadable_xlsx_fail_closed():
    with pytest.raises(HTTPException) as exc_info:
        IntakeService().parse_upload(
            filename="protected.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            file_bytes=b"not a workbook",
        )

    assert exc_info.value.status_code == 400
    assert exc_info.value.detail == {
        "code": "XLSX_PARSE_ERROR",
        "message": "XLSX file is unreadable or password-protected",
    }


@pytest.mark.parametrize("file_ext", ["csv", "xlsx"])
def test_parser_rejects_csv_and_xlsx_row_limit(file_ext: str):
    if file_ext == "csv":
        file_bytes = ("Sku Code\n" + "\n".join(f"SKU-{index}" for index in range(5001))).encode("utf-8")
    else:
        file_bytes = _xlsx_bytes(["Sku Code"], [[f"SKU-{index}"] for index in range(5001)])

    filename = f"too_many_rows.{file_ext}"

    _assert_parse_error(filename, file_bytes, "ROW_LIMIT_EXCEEDED")


@pytest.mark.parametrize(
    ("filename", "file_bytes"),
    [
        ("too_many_columns.csv", (",".join(f"H{index}" for index in range(101)) + "\n").encode("utf-8")),
        ("too_many_columns.xlsx", _xlsx_bytes([f"H{index}" for index in range(101)], [])),
    ],
)
def test_parser_rejects_csv_and_xlsx_column_limit(filename: str, file_bytes: bytes):
    _assert_parse_error(filename, file_bytes, "COLUMN_LIMIT_EXCEEDED")


@pytest.mark.parametrize(
    ("filename", "file_bytes"),
    [
        ("header_too_large.csv", (("H" * 256) + "\nvalue\n").encode("utf-8")),
        ("header_too_large.xlsx", _xlsx_bytes(["H" * 256], [["value"]])),
    ],
)
def test_parser_rejects_csv_and_xlsx_header_length(filename: str, file_bytes: bytes):
    _assert_parse_error(filename, file_bytes, "HEADER_TOO_LARGE")


@pytest.mark.parametrize(
    ("filename", "file_bytes"),
    [
        ("cell_too_large.csv", ("Name\n" + ("X" * 2001) + "\n").encode("utf-8")),
        ("cell_too_large.xlsx", _xlsx_bytes(["Name"], [["X" * 2001]])),
    ],
)
def test_parser_rejects_csv_and_xlsx_cell_length(filename: str, file_bytes: bytes):
    _assert_parse_error(filename, file_bytes, "CELL_TOO_LARGE")


@pytest.mark.asyncio
async def test_upload_mapping_validation_rows_and_issues_are_staging_only(async_session):
    await _ensure_intake_schema(async_session)

    async with _client_for(permissions=["intake:create"]) as create_client:
        create_response = await create_client.post(
            "/api/v1/intake/workspaces",
            json={"name": "U4-D staging flow", "source_type": "CUSTOMER_ONBOARDING"},
        )
    assert create_response.status_code == 201
    workspace_id = create_response.json()["data"]["workspace_id"]

    csv_payload = (
        "Sku Code,Name,Unit,Price,Extra\n"
        "SKU-1,Widget,,abc,keep\n"
        "SKU-1,,box,12.50,keep\n"
        ",Missing SKU,piece,1.00,keep\n"
        f"{'X' * 65},Too long,piece,2.00,keep\n"
    ).encode("utf-8")
    async with _client_for(permissions=["intake:update"]) as update_client:
        upload_response = await update_client.post(
            f"/api/v1/intake/workspaces/{workspace_id}/uploads",
            files={"file": ("products.csv", csv_payload, "text/csv")},
        )
        mapping_response = await update_client.put(
            f"/api/v1/intake/workspaces/{workspace_id}/mapping",
            json={"mapping": {"Sku Code": "sku_code", "Name": "name", "Unit": "unit", "Price": "unit_price"}},
        )
        validate_response = await update_client.post(f"/api/v1/intake/workspaces/{workspace_id}/validate")

    assert upload_response.status_code == 201
    upload_data = upload_response.json()["data"]
    assert upload_data["row_count"] == 4
    assert upload_data["headers_normalized"]["Sku Code"] == "sku_code"

    assert mapping_response.status_code == 200
    mapping_data = mapping_response.json()["data"]
    assert mapping_data["mapped_rows"] == 4
    assert mapping_data["unit_default_note"] == "Missing unit is documented for review; U4-D does not mutate it to a default value."

    assert validate_response.status_code == 200
    validation_data = validate_response.json()["data"]
    assert validation_data["status"] == "NEEDS_REVIEW"
    assert validation_data["error_count"] >= 5
    assert validation_data["warning_count"] >= 2

    async with _client_for(permissions=["intake:read"]) as read_client:
        rows_response = await read_client.get(f"/api/v1/intake/workspaces/{workspace_id}/rows")
        issues_response = await read_client.get(f"/api/v1/intake/workspaces/{workspace_id}/issues")

    assert rows_response.status_code == 200
    rows = rows_response.json()["data"]["items"]
    assert rows[0]["source_row_number"] == 2
    assert rows[0]["raw_values"]["Extra"] == "keep"
    assert rows[0]["normalized_values"] == {"sku_code": "SKU-1", "name": "Widget", "unit": "", "unit_price": "abc"}

    assert issues_response.status_code == 200
    issue_codes = {issue["code"] for issue in issues_response.json()["data"]["items"]}
    assert {
        "DUPLICATE_STAGED_SKU_CODE",
        "FIELD_TOO_LONG",
        "INVALID_UNIT_PRICE",
        "MISSING_NAME",
        "MISSING_SKU_CODE",
        "UNIT_DEFAULT_AVAILABLE",
        "UNMAPPED_EXTRA_COLUMN",
    }.issubset(issue_codes)


@pytest.mark.asyncio
async def test_upload_requires_create_or_update_permission(async_session):
    await _ensure_intake_schema(async_session)

    async with _client_for(permissions=["intake:create"]) as create_client:
        create_response = await create_client.post(
            "/api/v1/intake/workspaces",
            json={"name": "Upload auth", "source_type": "CUSTOMER_ONBOARDING"},
        )
    workspace_id = create_response.json()["data"]["workspace_id"]

    async with _client_for(permissions=["intake:read"]) as read_client:
        response = await read_client.post(
            f"/api/v1/intake/workspaces/{workspace_id}/uploads",
            files={"file": ("products.csv", b"Sku Code,Name\nSKU-1,Name\n", "text/csv")},
        )

    assert response.status_code == 403
    assert response.json()["detail"]["code"] == "PERMISSION_DENIED"


def test_u4d_implementation_has_no_sku_public_token_or_u3_import_surface():
    source = "\n".join(
        [
            (BACKEND_DIR / "api" / "v1" / "intake.py").read_text(encoding="utf-8"),
            (BACKEND_DIR / "services" / "intake_service.py").read_text(encoding="utf-8"),
        ]
    )

    forbidden = [
        "intake_public_tokens",
        "intake_exports",
        "ImportService",
        "sku_import",
        "skus/import",
        "SKU(",
    ]
    for value in forbidden:
        assert value not in source, f"Forbidden U4-D implementation surface found: {value}"
