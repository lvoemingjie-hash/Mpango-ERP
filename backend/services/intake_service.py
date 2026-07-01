"""U4-D data intake parser, mapping, and validation service."""
from __future__ import annotations

import csv
import hashlib
import io
import re
import uuid
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable, Optional
from zipfile import BadZipFile

from fastapi import HTTPException, status
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from models.intake import IntakeProductRow, IntakeUpload, IntakeValidationIssue, IntakeWorkspace


ALLOWED_FILE_EXTENSIONS = {"csv", "xlsx"}
MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_ROWS = 5000
MAX_COLUMNS = 100
MAX_CELL_LENGTH = 2000
MAX_HEADER_LENGTH = 255
TARGET_FIELDS = {"sku_code", "name", "unit", "category", "unit_price", "barcode"}
FIELD_LENGTHS = {
    "sku_code": 64,
    "name": 255,
    "unit": 32,
    "category": 64,
    "barcode": 128,
}


@dataclass(frozen=True)
class ParsedIntakeFile:
    filename: str
    content_type: Optional[str]
    file_ext: str
    file_size_bytes: int
    sha256: str
    headers_raw: list[str]
    headers_normalized: dict[str, str]
    rows: list[dict[str, Any]]
    source_row_numbers: list[int]
    parse_summary: dict[str, Any]


class IntakeService:
    """Service for U4 staging-only parser/preview operations."""

    def parse_upload(self, *, filename: str, content_type: Optional[str], file_bytes: bytes) -> ParsedIntakeFile:
        if not file_bytes:
            self._raise_bad_request("EMPTY_FILE", "Uploaded file is empty")
        if len(file_bytes) > MAX_UPLOAD_BYTES:
            self._raise_bad_request("FILE_TOO_LARGE", "Uploaded file exceeds the 10 MB limit")

        file_ext = Path(filename or "").suffix.lower().lstrip(".")
        if file_ext not in ALLOWED_FILE_EXTENSIONS:
            self._raise_bad_request("UNSUPPORTED_FILE_TYPE", "Only CSV and XLSX uploads are supported")

        if file_ext == "csv":
            headers_raw, rows, source_row_numbers, summary = self._parse_csv(file_bytes)
        else:
            headers_raw, rows, source_row_numbers, summary = self._parse_xlsx(file_bytes)

        headers_normalized = self.normalize_headers(headers_raw)
        return ParsedIntakeFile(
            filename=filename or f"upload.{file_ext}",
            content_type=content_type,
            file_ext=file_ext,
            file_size_bytes=len(file_bytes),
            sha256=hashlib.sha256(file_bytes).hexdigest(),
            headers_raw=headers_raw,
            headers_normalized=headers_normalized,
            rows=rows,
            source_row_numbers=source_row_numbers,
            parse_summary=summary,
        )

    async def create_upload_rows(
        self,
        db: AsyncSession,
        *,
        tenant_id: uuid.UUID,
        workspace: IntakeWorkspace,
        parsed: ParsedIntakeFile,
        user_id: Optional[uuid.UUID],
    ) -> tuple[IntakeUpload, int]:
        upload = IntakeUpload(
            tenant_id=tenant_id,
            workspace_id=workspace.id,
            filename=parsed.filename,
            content_type=parsed.content_type,
            file_ext=parsed.file_ext,
            file_size_bytes=parsed.file_size_bytes,
            sha256=parsed.sha256,
            status="PARSED",
            row_count=len(parsed.rows),
            column_count=len(parsed.headers_raw),
            headers_raw=parsed.headers_raw,
            headers_normalized=parsed.headers_normalized,
            parse_summary=parsed.parse_summary,
            created_by=user_id,
            updated_by=user_id,
        )
        db.add(upload)
        await db.flush()

        for index, raw_values in enumerate(parsed.rows):
            db.add(
                IntakeProductRow(
                    tenant_id=tenant_id,
                    workspace_id=workspace.id,
                    upload_id=upload.id,
                    source_row_number=parsed.source_row_numbers[index],
                    row_index=index,
                    raw_values=raw_values,
                    normalized_values={},
                    review_status="UNREVIEWED",
                    created_by=user_id,
                    updated_by=user_id,
                )
            )

        workspace.status = "UPLOADED"
        workspace.updated_by = user_id
        await db.flush()
        return upload, len(parsed.rows)

    async def apply_mapping(
        self,
        db: AsyncSession,
        *,
        tenant_id: uuid.UUID,
        workspace: IntakeWorkspace,
        mapping: dict[str, str],
        user_id: Optional[uuid.UUID],
    ) -> int:
        self._validate_mapping(mapping)
        rows = (
            await db.execute(
                select(IntakeProductRow).where(
                    IntakeProductRow.tenant_id == tenant_id,
                    IntakeProductRow.workspace_id == workspace.id,
                    IntakeProductRow.is_deleted.is_(False),
                )
            )
        ).scalars().all()

        upload_ids = {row.upload_id for row in rows}
        uploads = []
        if upload_ids:
            uploads = (
                await db.execute(
                    select(IntakeUpload).where(
                        IntakeUpload.tenant_id == tenant_id,
                        IntakeUpload.workspace_id == workspace.id,
                        IntakeUpload.id.in_(upload_ids),
                        IntakeUpload.is_deleted.is_(False),
                    )
                )
            ).scalars().all()
        normalized_header_maps = {upload.id: upload.headers_normalized or {} for upload in uploads}

        current_version = max((row.mapping_version for row in rows), default=0)
        next_version = current_version + 1
        for row in rows:
            header_map = normalized_header_maps.get(row.upload_id, {})
            row.normalized_values = self._build_normalized_values(
                raw_values=row.raw_values or {},
                header_map=header_map,
                mapping=mapping,
            )
            row.mapping_version = next_version
            self._copy_known_fields(row)
            row.updated_by = user_id

        metadata = dict(workspace.metadata_json or {})
        metadata["column_mapping"] = mapping
        metadata["unit_default_note"] = "Missing unit is documented for review; U4-D does not mutate it to a default value."
        workspace.metadata_json = metadata
        workspace.status = "MAPPED"
        workspace.updated_by = user_id
        await db.flush()
        return len(rows)

    async def validate_workspace(
        self,
        db: AsyncSession,
        *,
        tenant_id: uuid.UUID,
        workspace: IntakeWorkspace,
        user_id: Optional[uuid.UUID],
    ) -> dict[str, int | str]:
        await db.execute(
            delete(IntakeValidationIssue).where(
                IntakeValidationIssue.tenant_id == tenant_id,
                IntakeValidationIssue.workspace_id == workspace.id,
            )
        )

        rows = (
            await db.execute(
                select(IntakeProductRow).where(
                    IntakeProductRow.tenant_id == tenant_id,
                    IntakeProductRow.workspace_id == workspace.id,
                    IntakeProductRow.is_deleted.is_(False),
                ).order_by(IntakeProductRow.upload_id, IntakeProductRow.row_index)
            )
        ).scalars().all()

        uploads = (
            await db.execute(
                select(IntakeUpload).where(
                    IntakeUpload.tenant_id == tenant_id,
                    IntakeUpload.workspace_id == workspace.id,
                    IntakeUpload.is_deleted.is_(False),
                )
            )
        ).scalars().all()
        mapping = dict((workspace.metadata_json or {}).get("column_mapping") or {})
        upload_header_maps = {upload.id: upload.headers_normalized or {} for upload in uploads}

        issues: list[IntakeValidationIssue] = []
        sku_rows: dict[str, list[IntakeProductRow]] = {}

        for row in rows:
            values = row.normalized_values or {}
            sku_code = self._string_or_none(values.get("sku_code"))
            name = self._string_or_none(values.get("name"))

            if not sku_code:
                issues.append(self._issue(tenant_id, workspace.id, row, "ERROR", "MISSING_SKU_CODE", "sku_code", "Missing sku_code", True))
            else:
                sku_rows.setdefault(sku_code, []).append(row)
            if not name:
                issues.append(self._issue(tenant_id, workspace.id, row, "ERROR", "MISSING_NAME", "name", "Missing name", True))

            unit_price = self._string_or_none(values.get("unit_price"))
            if unit_price is not None and self._decimal_or_none(unit_price) is None:
                issues.append(
                    self._issue(
                        tenant_id,
                        workspace.id,
                        row,
                        "ERROR",
                        "INVALID_UNIT_PRICE",
                        "unit_price",
                        "unit_price must be a valid decimal value",
                        True,
                    )
                )

            for field, max_length in FIELD_LENGTHS.items():
                value = self._string_or_none(values.get(field))
                if value and len(value) > max_length:
                    issues.append(
                        self._issue(
                            tenant_id,
                            workspace.id,
                            row,
                            "ERROR",
                            "FIELD_TOO_LONG",
                            field,
                            f"{field} exceeds {max_length} characters",
                            True,
                        )
                    )

            unit = self._string_or_none(values.get("unit"))
            if not unit:
                issues.append(
                    self._issue(
                        tenant_id,
                        workspace.id,
                        row,
                        "WARNING",
                        "UNIT_DEFAULT_AVAILABLE",
                        "unit",
                        "Unit is optional; no default was written to staged data.",
                        False,
                    )
                )

        for sku_code, duplicate_rows in sku_rows.items():
            if len(duplicate_rows) < 2:
                continue
            for row in duplicate_rows:
                issues.append(
                    self._issue(
                        tenant_id,
                        workspace.id,
                        row,
                        "ERROR",
                        "DUPLICATE_STAGED_SKU_CODE",
                        "sku_code",
                        f"Duplicate staged sku_code '{sku_code}' in this workspace",
                        True,
                    )
                )

        mapped_sources = {self.normalize_header(source) for source in mapping}
        for upload in uploads:
            header_map = upload_header_maps.get(upload.id, {})
            unmapped_headers = [raw for raw, normalized in header_map.items() if normalized not in mapped_sources]
            for raw_header in unmapped_headers:
                issues.append(
                    IntakeValidationIssue(
                        tenant_id=tenant_id,
                        workspace_id=workspace.id,
                        upload_id=upload.id,
                        row_id=None,
                        source_row_number=None,
                        severity="WARNING",
                        code="UNMAPPED_EXTRA_COLUMN",
                        field=None,
                        source_header=raw_header,
                        message=f"Column '{raw_header}' is not mapped and will remain staged only.",
                        is_blocking=False,
                    )
                )

        for issue in issues:
            db.add(issue)

        error_count = sum(1 for issue in issues if issue.severity == "ERROR")
        warning_count = sum(1 for issue in issues if issue.severity == "WARNING")
        workspace.status = "NEEDS_REVIEW" if error_count else "READY_FOR_EXPORT"
        workspace.updated_by = user_id
        await db.flush()
        return {
            "status": workspace.status,
            "row_count": len(rows),
            "error_count": error_count,
            "warning_count": warning_count,
        }

    def normalize_headers(self, headers: Iterable[str]) -> dict[str, str]:
        header_list = list(headers)
        used: dict[str, int] = {}
        normalized: dict[str, str] = {}
        for raw_header, header_key in zip(header_list, self._header_keys(header_list)):
            header = str(raw_header or "")
            base = self.normalize_header(header)
            count = used.get(base, 0)
            used[base] = count + 1
            normalized[header_key] = base if count == 0 else f"{base}_{count + 1}"
        return normalized

    def normalize_header(self, header: str) -> str:
        value = str(header or "").strip().lower()
        value = re.sub(r"[^a-z0-9]+", "_", value)
        value = re.sub(r"_+", "_", value).strip("_")
        return value or "column"

    def _parse_csv(self, file_bytes: bytes) -> tuple[list[str], list[dict[str, Any]], list[int], dict[str, Any]]:
        try:
            text = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"code": "ENCODING_ERROR", "message": "CSV must be UTF-8 or UTF-8-sig encoded"},
            ) from exc

        sample = text[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample) if sample else csv.excel
        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(io.StringIO(text), dialect)
        rows = list(reader)
        if not rows:
            self._raise_bad_request("EMPTY_FILE", "CSV has no header row")
        header_index = self._first_non_empty_row_index(rows)
        if header_index is None:
            self._raise_bad_request("EMPTY_FILE", "CSV has no non-empty rows")
        headers = [str(value or "").strip() for value in rows[header_index]]
        self._validate_headers(headers)

        parsed_rows: list[dict[str, Any]] = []
        source_row_numbers: list[int] = []
        for row_number, values in enumerate(rows[header_index + 1 :], start=header_index + 2):
            if self._row_is_empty(values):
                continue
            self._validate_data_row(values)
            parsed_rows.append(self._row_dict(headers, values))
            source_row_numbers.append(row_number)
            self._validate_row_count(len(parsed_rows))

        return headers, parsed_rows, source_row_numbers, {"parser": "csv", "sheet_name": None, "header_row_number": header_index + 1}

    def _parse_xlsx(self, file_bytes: bytes) -> tuple[list[str], list[dict[str, Any]], list[int], dict[str, Any]]:
        try:
            from openpyxl import load_workbook
            from openpyxl.utils.exceptions import InvalidFileException
        except ImportError as exc:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail={"code": "XLSX_PARSER_UNAVAILABLE", "message": "XLSX parser dependency is not installed"},
            ) from exc

        try:
            workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except (InvalidFileException, OSError, RuntimeError, ValueError, KeyError, BadZipFile) as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"code": "XLSX_PARSE_ERROR", "message": "XLSX file is unreadable or password-protected"},
            ) from exc

        try:
            for worksheet in workbook.worksheets:
                rows = list(worksheet.iter_rows(values_only=True))
                header_index = self._first_non_empty_row_index(rows)
                if header_index is None:
                    continue
                headers = [self._cell_to_string(value).strip() for value in rows[header_index]]
                self._validate_headers(headers)
                parsed_rows: list[dict[str, Any]] = []
                source_row_numbers: list[int] = []
                for row_number, values in enumerate(rows[header_index + 1 :], start=header_index + 2):
                    if self._row_is_empty(values):
                        continue
                    self._validate_data_row(values)
                    parsed_rows.append(self._row_dict(headers, list(values)))
                    source_row_numbers.append(row_number)
                    self._validate_row_count(len(parsed_rows))
                return headers, parsed_rows, source_row_numbers, {
                    "parser": "xlsx",
                    "sheet_name": worksheet.title,
                    "header_row_number": header_index + 1,
                }
        finally:
            workbook.close()

        self._raise_bad_request("EMPTY_FILE", "XLSX has no non-empty sheets")

    def _row_dict(self, headers: list[str], values: list[Any]) -> dict[str, Any]:
        row: dict[str, Any] = {}
        for index, header in enumerate(self._header_keys(headers)):
            value = values[index] if index < len(values) else None
            row[header] = self._cell_to_string(value)
        return row

    def _header_keys(self, headers: Iterable[str]) -> list[str]:
        used: dict[str, int] = {}
        keys: list[str] = []
        for header in headers:
            value = str(header or "")
            count = used.get(value, 0)
            used[value] = count + 1
            keys.append(value if count == 0 else f"{value}__{count + 1}")
        return keys

    def _cell_to_string(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, Decimal):
            return format(value, "f")
        return str(value).strip()

    def _first_non_empty_row_index(self, rows: list[Any]) -> Optional[int]:
        for index, row in enumerate(rows):
            if not self._row_is_empty(row):
                return index
        return None

    def _row_is_empty(self, row: Any) -> bool:
        return all(self._cell_to_string(value) == "" for value in row)

    def _validate_headers(self, headers: list[str]) -> None:
        if not headers or all(not header for header in headers):
            self._raise_bad_request("EMPTY_HEADER", "Upload must include a non-empty header row")
        if len(headers) > MAX_COLUMNS:
            self._raise_bad_request("COLUMN_LIMIT_EXCEEDED", f"Upload exceeds the {MAX_COLUMNS} column limit")
        for header in headers:
            if len(header) > MAX_HEADER_LENGTH:
                self._raise_bad_request("HEADER_TOO_LARGE", f"Header exceeds the {MAX_HEADER_LENGTH} character limit")

    def _validate_data_row(self, values: Iterable[Any]) -> None:
        value_list = list(values)
        if len(value_list) > MAX_COLUMNS:
            self._raise_bad_request("COLUMN_LIMIT_EXCEEDED", f"Upload exceeds the {MAX_COLUMNS} column limit")
        for value in value_list:
            if len(self._cell_to_string(value)) > MAX_CELL_LENGTH:
                self._raise_bad_request("CELL_TOO_LARGE", f"Cell exceeds the {MAX_CELL_LENGTH} character limit")

    def _validate_row_count(self, row_count: int) -> None:
        if row_count > MAX_ROWS:
            self._raise_bad_request("ROW_LIMIT_EXCEEDED", f"Upload exceeds the {MAX_ROWS} row limit")

    def _validate_mapping(self, mapping: dict[str, str]) -> None:
        if not mapping:
            self._raise_bad_request("INVALID_MAPPING", "Mapping must include at least one column")
        invalid_targets = sorted({target for target in mapping.values() if target not in TARGET_FIELDS})
        if invalid_targets:
            self._raise_bad_request("INVALID_MAPPING", f"Unsupported target fields: {', '.join(invalid_targets)}")

    def _build_normalized_values(self, *, raw_values: dict[str, Any], header_map: dict[str, str], mapping: dict[str, str]) -> dict[str, Any]:
        normalized_values: dict[str, Any] = {}
        source_lookup = {self.normalize_header(source): target for source, target in mapping.items()}
        for raw_header, raw_value in raw_values.items():
            normalized_source = header_map.get(raw_header, self.normalize_header(raw_header))
            target = source_lookup.get(normalized_source)
            if target:
                normalized_values[target] = raw_value
        return normalized_values

    def _copy_known_fields(self, row: IntakeProductRow) -> None:
        values = row.normalized_values or {}
        row.sku_code = self._bounded_string_or_none(values.get("sku_code"), FIELD_LENGTHS["sku_code"])
        row.name = self._bounded_string_or_none(values.get("name"), FIELD_LENGTHS["name"])
        row.unit = self._bounded_string_or_none(values.get("unit"), FIELD_LENGTHS["unit"])
        row.category = self._bounded_string_or_none(values.get("category"), FIELD_LENGTHS["category"])
        row.barcode = self._bounded_string_or_none(values.get("barcode"), FIELD_LENGTHS["barcode"])
        row.unit_price = self._decimal_or_none(values.get("unit_price"))

    def _decimal_or_none(self, value: Any) -> Optional[Decimal]:
        string_value = self._string_or_none(value)
        if string_value is None:
            return None
        try:
            return Decimal(string_value)
        except InvalidOperation:
            return None

    def _string_or_none(self, value: Any) -> Optional[str]:
        string_value = self._cell_to_string(value)
        return string_value if string_value else None

    def _bounded_string_or_none(self, value: Any, max_length: int) -> Optional[str]:
        string_value = self._string_or_none(value)
        if string_value is None or len(string_value) > max_length:
            return None
        return string_value

    def _issue(
        self,
        tenant_id: uuid.UUID,
        workspace_id: uuid.UUID,
        row: IntakeProductRow,
        severity: str,
        code: str,
        field: Optional[str],
        message: str,
        is_blocking: bool,
    ) -> IntakeValidationIssue:
        return IntakeValidationIssue(
            tenant_id=tenant_id,
            workspace_id=workspace_id,
            upload_id=row.upload_id,
            row_id=row.id,
            source_row_number=row.source_row_number,
            severity=severity,
            code=code,
            field=field,
            source_header=None,
            message=message,
            is_blocking=is_blocking,
        )

    def _raise_bad_request(self, code: str, message: str) -> None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail={"code": code, "message": message})
